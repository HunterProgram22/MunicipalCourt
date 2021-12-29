
import os
import sys
import pathlib
from dataclasses import dataclass, field, asdict
from loguru import logger

from openpyxl import load_workbook
from docxtpl import DocxTemplate


PATH = str(pathlib.Path().absolute())
TEMPLATE_PATH = "C:\\Users\\jkudela\\OneDrive - delawareoh\\Desktop\\Benefits\\Template_Benefits.docx"
SAVE_PATH = "C:\\Users\\jkudela\\OneDrive - delawareoh\\Desktop\\Benefits\\Employees\\"

def create_entry(data):
    doc = DocxTemplate(TEMPLATE_PATH)
    doc.render(data.get_case_information())
    docname = set_document_name(data)
    doc.save(SAVE_PATH + docname)
    #os.startfile(SAVE_PATH + docname)


def set_document_name(data):
    docname = data.name + ".docx"
    return docname


def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file, data_only=True)
    worksheet = workbook.active
    row_count = 25
    for row in range(3, row_count):
        name = worksheet.cell(row=row, column=3)
        hire_date = worksheet.cell(row=row, column=4)
        position = worksheet.cell(row=row, column=5)
        position_pay_grade = worksheet.cell(row=row, column=6)
        position_pay_range_min = worksheet.cell(row=row, column=7)
        position_pay_range_max = worksheet.cell(row=row, column=8)
        pay_rate_2021 = worksheet.cell(row=row, column=9)
        pay_rate_2022 = worksheet.cell(row=row, column=10)
        prior_service_verification = worksheet.cell(row=row, column=11)
        vacation_accrural_rate_2021 = worksheet.cell(row=row, column=12)
        vacation_accrural_rate_2022 = worksheet.cell(row=row, column=13)
        longevity_pay = worksheet.cell(row=row, column=14)
        total_compensation_2022 = worksheet.cell(row=row, column=15)
        percent_increase = worksheet.cell(row=row, column=16)
        city_service = worksheet.cell(row=row, column=17)
        employee = EmployeeInformation(
                name.value,
                hire_date.value,
                position.value,
                position_pay_grade.value,
                position_pay_range_min.value,
                position_pay_range_max.value,
                pay_rate_2021.value,
                pay_rate_2022.value,
                prior_service_verification.value,
                vacation_accrural_rate_2021.value,
                vacation_accrural_rate_2022.value,
                longevity_pay.value,
                total_compensation_2022.value,
                percent_increase.value,
                city_service.value,
                )
        data.append(employee)
    return data

@dataclass
class EmployeeInformation:
    name: str  = None
    hire_date: str = None
    position: str = None
    position_pay_grade: str = None
    position_pay_range_min: str = None
    position_pay_range_max: str = None
    pay_rate_2021: float = None
    pay_rate_2022: float = None
    prior_service_verification: str = None
    vacation_accrural_rate_2021: float = None
    vacation_accrural_rate_2022: float = None
    longevity_pay: float = None
    total_compensation_2022: float = None
    percent_increase: float = None
    city_service: float = None

    def get_case_information(self):
        """Returns a dictionary with all of the employee information required
        to populate an entry."""
        return asdict(self)

data_for_entries = return_data_from_excel("C:\\Users\\jkudela\\OneDrive - delawareoh\\Desktop\\Benefits\\BenefitsSummary.xlsx")

print(data_for_entries[0])
for index, person in enumerate(data_for_entries):
    create_entry(data_for_entries[index])
