"""Scripts for calculating the total amount of costs and fines that would be
due for a defendant during 2nd Chance Week."""
import os
import sys
import pathlib
from dataclasses import dataclass, field, asdict

from openpyxl import load_workbook

PATH = str(pathlib.Path().absolute())

DOC_PATH = 'C:\\Users\\jkudela\\OneDrive - delawareoh\\Desktop\\Copy of Amnesty Heard Jan 2018.xlsx'


def return_data_from_excel(excel_file):
    data = []
    workbook = load_workbook(excel_file, data_only=True)
    worksheet = workbook.active
    row_count = 25
    for row in range(3, worksheet.max_row):
        case_number = worksheet.cell(row=row, column=2).value[:-2]
        total_balance_due = worksheet.cell(row=row, column=4).value
        charge_code = worksheet.cell(row=row, column=6).value
        initial_amount_owed = worksheet.cell(row=row, column=7).value
        current_balance_due = worksheet.cell(row=row, column=8).value
        data.append((
            case_number,
            total_balance_due,
            charge_code,
            initial_amount_owed,
            current_balance_due,
        ))
    return data


@dataclass
class ChargeCode:
    charge_code: str
    initial_amount_owed: float
    remaining_balance_owed: float


@dataclass
class Case:
   case_number: str
   total_balance_due: float
   amnesty_balance_due: float

   def calculate_totals(self):
       print("calculating")

case_number_list = []

for item in return_data_from_excel(DOC_PATH):
    case_number = item[0]
    total_balance_due = item[1]
    if case_number in case_number_list:
        print(case_number)
        print(total_balance_due)
    else:
        new_case = Case(item[0], item[1], 0)
        case_number_list.append(new_case.case_number)
        new_case.calculate_totals()

print(case_number_list)